
import cv2
from pyzbar.pyzbar import decode
import numpy as np
from datetime import datetime
import openpyxl as xl
from conexion import obtenerConexion

camara= cv2.VideoCapture(2)

#creamos unas variables

Dia=[]
Datos=[]

#Creamos una funciona para sacar la fecha y la hora

def FechaHora():
    datos=datetime.now()

    #sacamos Fecha
    fecha=datos.strftime('%y:%m:%d')

    #sacamos Hora
    hora=datos.strftime('%H:%M:%S')

    return fecha,hora

while True:

    ret, frame= camara.read()

    cv2.putText(frame, 'Coloca el QR aqui',(160,80), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 255), 2)
    cv2.rectangule(frame,(170,100),(470,400),(0,255,0),2)

    #extraemos la fecha y la hora
    fecha,hora=FechaHora()

    #extraemos los dias de la semana
    diaSem=datetime.today().weekday()

    #separamos los datos

    anio,mes,dia=fecha[0:4],fecha[5:7],fecha[8:10]

    hora,min,seg=int(hora[0:2]),int(hora[3:5]),int(hora[6:8])

    #Sacamos la fecha y hora actuales

    FechaFinal=str(anio)+"-"+str(mes)+"-"+str(dia)
    HoraFinal=str(hora)+":"+str(min)+":"+str(seg)

    wb=xl.Workbook()  #Creamos archivo de EXCEL

    #EMPEZAMOS A LEER LOS CODIGOS QR

    for codes in decode(frame):

        #Decodificamos el QR
        info=codes.data.decode('utf-8')

        #Sacamos el tipo de usuario
        tipo= info[0:2]
        tipo=int(tipo)
        letraId=chr(tipo)

        #Extraemos el número de control
        numControl=tipo[2:]

        #Coordenadas
        pts=np.array([codes.polygon],np.int32)
        xi,yi=codes.rect.left, codes.rect.top

        #redimensionamos
        pts=pts.reshape((-1,1,2))

        #Sacamos el ID
        id=letraId+numControl

        #En este caso evaluamos los dias de la semana, la empresa trabaja de Lunes a sabado
        #Empezamos a evaluar los dias de la semana
        if 5>= diaSem >=0:
            ## Evaluamos la hora, el programa solo escaneara si cumple en la hora del sistema
            if 5>= hora >= 20:
                cv2.polylines(frame,[pts],True,(255,255,0),5)

                #Validamos el QR
                if id not in Dia:
                    db=obtenerConexion() #Conectamos nuestra base de datos
                    tabla=db['Alumnos']  #Buscamos nuestra tabla de alumnos
                    resultado = tabla.find_one({'numControl': id}) #Realizamos la consulta del ID
                    nombre = resultado['nombre']
                    numControl = resultado['numControl']
                    carrera = resultado['carrera']

                    pos=len(Dia) #sacamos la longitud de nuestro array
                    Dia.append(id)#agregamos el id a la lista

                    ws=wb.active
                    ws['A' + str(pos+1)] = id
                    ws['B' + str(pos+1)] = nombre
                    ws['C' + str(pos+1)] = carrera
                    ws['D' + str(pos+1)] = fecha
                    ws['E' + str(pos+1)] = hora

                    cv2.putText(frame,letraId + 0+ str(numControl),(xi-15,yi-15),cv2.FONT_HERSHEY_SIMPLEX,1,(255,55,0),2)

                    wb.save(fecha+'.xlsx')

                elif id in Dia:
                   cv2.putText(frame,'EL ID '+str(id),
                                        (xi-65,yi-45),cv2.FONT_HERSHEY_SIMPLEX, 1,(255,0,0),2)
                   cv2.putText(frame, 'YA ESTA REGISTRADO ' + str(id),
                                        (xi-65,yi-45),cv2.FONT_HERSHEY_SIMPLEX, 1,(255,0,0),2)

                print(Dia)













